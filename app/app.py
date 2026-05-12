import streamlit as st
import pandas as pd
import tempfile

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# =====================================================
# PAGE CONFIG
# =====================================================

st.set_page_config(
    page_title="Retail UID Difference Checker",
    layout="wide"
)

st.title("🔍 Retail UID Difference Checker")

# =====================================================
# HELPERS
# =====================================================

@st.cache_data
def load_file(file):

    if file.name.endswith(".csv"):
        df = pd.read_csv(file, dtype=str)

    else:
        df = pd.read_excel(
            file,
            engine="openpyxl",
            dtype=str
        )

    df.columns = df.columns.str.strip()

    return df


def clean_upc(series):

    return (
        series.astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.replace(r"\D", "", regex=True)
        .str.strip()
    )


def find_column(columns, possible_names):

    for possible in possible_names:

        for col in columns:

            if possible.lower() in col.lower():
                return col

    return columns[0]


# =====================================================
# FILE UPLOADS
# =====================================================

st.header("Upload Files")

master_file = st.file_uploader(
    "Upload Master Product File",
    type=["xlsx", "csv"]
)

bu_file = st.file_uploader(
    "Upload BU Product File",
    type=["xlsx", "csv"]
)

# =====================================================
# MAIN PROCESS
# =====================================================

if master_file and bu_file:

    master_df = load_file(master_file)
    bu_df = load_file(bu_file)

    st.success("Files Loaded")

    # =====================================================
    # AUTO DETECT COLUMNS
    # =====================================================

    master_upc_default = find_column(
        master_df.columns,
        ["upc", "productupc"]
    )

    master_uid_default = find_column(
        master_df.columns,
        ["item no", "productid", "retail uid"]
    )

    bu_product_upc_default = find_column(
        bu_df.columns,
        ["productupc", "product upc"]
    )

    bu_unit_upc_default = find_column(
        bu_df.columns,
        ["unitupc", "unit upc"]
    )

    bu_uid_default = find_column(
        bu_df.columns,
        ["productid", "item no", "retail uid"]
    )

    # =====================================================
    # COLUMN MAPPING UI
    # =====================================================

    st.header("Map Columns")

    master_upc_col = st.selectbox(
        "Master UPC Column",
        master_df.columns,
        index=list(master_df.columns).index(
            master_upc_default
        )
    )

    master_uid_col = st.selectbox(
        "Master Retail UID Column",
        master_df.columns,
        index=list(master_df.columns).index(
            master_uid_default
        )
    )

    bu_product_upc_col = st.selectbox(
        "BU ProductUPC Column",
        bu_df.columns,
        index=list(bu_df.columns).index(
            bu_product_upc_default
        )
    )

    bu_unit_upc_col = st.selectbox(
        "BU UnitUPC Column",
        bu_df.columns,
        index=list(bu_df.columns).index(
            bu_unit_upc_default
        )
    )

    bu_uid_col = st.selectbox(
        "BU Retail UID Column",
        bu_df.columns,
        index=list(bu_df.columns).index(
            bu_uid_default
        )
    )

    # =====================================================
    # PROCESS BUTTON
    # =====================================================

    if st.button("🚀 Compare Files"):

        progress = st.progress(0)
        status = st.empty()

        with st.spinner("Processing files..."):

            # =====================================================
            # CLEAN UPCS
            # =====================================================

            status.text("Cleaning UPC columns...")
            progress.progress(15)

            master_df["MASTER_UPC_CLEAN"] = clean_upc(
                master_df[master_upc_col]
            )

            bu_df["PRODUCT_UPC_CLEAN"] = clean_upc(
                bu_df[bu_product_upc_col]
            )

            bu_df["UNIT_UPC_CLEAN"] = clean_upc(
                bu_df[bu_unit_upc_col]
            )

            # =====================================================
            # BUILD MASTER LOOKUP
            # =====================================================

            status.text("Building UPC lookup...")
            progress.progress(35)

            master_lookup = (
                master_df[
                    [
                        "MASTER_UPC_CLEAN",
                        master_uid_col
                    ]
                ]
                .dropna(subset=["MASTER_UPC_CLEAN"])
                .drop_duplicates(subset=["MASTER_UPC_CLEAN"])
                .rename(columns={
                    master_uid_col: "MASTER_UID"
                })
            )

            master_lookup_dict = dict(
                zip(
                    master_lookup["MASTER_UPC_CLEAN"],
                    master_lookup["MASTER_UID"]
                )
            )

            # =====================================================
            # MATCH LOGIC
            # =====================================================

            status.text("Matching UPCs...")
            progress.progress(60)

            def lookup_master_uid(row):

                product_upc = row["PRODUCT_UPC_CLEAN"]
                unit_upc = row["UNIT_UPC_CLEAN"]

                # TRY PRODUCT UPC FIRST
                if product_upc in master_lookup_dict:

                    return (
                        master_lookup_dict[product_upc],
                        "ProductUPC"
                    )

                # FALLBACK TO UNIT UPC
                if unit_upc in master_lookup_dict:

                    return (
                        master_lookup_dict[unit_upc],
                        "UnitUPC"
                    )

                return (None, "No Match")

            lookup_results = bu_df.apply(
                lookup_master_uid,
                axis=1
            )

            bu_df["MASTER_UID"] = lookup_results.apply(
                lambda x: x[0]
            )

            bu_df["MATCH_SOURCE"] = lookup_results.apply(
                lambda x: x[1]
            )

            # =====================================================
            # FLAG MISMATCHES
            # =====================================================

            status.text("Finding UID mismatches...")
            progress.progress(80)

            flagged = bu_df[
                (
                    bu_df["MASTER_UID"].notna()
                )
                &
                (
                    bu_df[bu_uid_col]
                    .astype(str)
                    .str.strip()
                    !=
                    bu_df["MASTER_UID"]
                    .astype(str)
                    .str.strip()
                )
            ].copy()

            flagged["BU_UID"] = flagged[bu_uid_col]

            # =====================================================
            # SUMMARY
            # =====================================================

            total_rows = len(bu_df)

            mismatch_count = len(flagged)

            matched_product_upc = len(
                bu_df[
                    bu_df["MATCH_SOURCE"] == "ProductUPC"
                ]
            )

            matched_unit_upc = len(
                bu_df[
                    bu_df["MATCH_SOURCE"] == "UnitUPC"
                ]
            )

            no_match_count = len(
                bu_df[
                    bu_df["MATCH_SOURCE"] == "No Match"
                ]
            )

            summary_df = pd.DataFrame({
                "Metric": [
                    "Total Rows",
                    "UID Mismatches",
                    "Matched by ProductUPC",
                    "Matched by UnitUPC",
                    "No Match"
                ],
                "Value": [
                    total_rows,
                    mismatch_count,
                    matched_product_upc,
                    matched_unit_upc,
                    no_match_count
                ]
            })

            # =====================================================
            # OUTPUT COLUMNS
            # =====================================================

            output_columns = [
                bu_product_upc_col,
                bu_unit_upc_col,
                "BU_UID",
                "MASTER_UID",
                "MATCH_SOURCE"
            ]

            remaining_cols = [
                c for c in flagged.columns
                if c not in output_columns
                and c not in [
                    "PRODUCT_UPC_CLEAN",
                    "UNIT_UPC_CLEAN"
                ]
            ]

            flagged = flagged[
                output_columns + remaining_cols
            ]

            progress.progress(100)
            status.text("Done!")

        # =====================================================
        # DISPLAY RESULTS
        # =====================================================

        st.success(
            f"Found {len(flagged)} Retail UID mismatches"
        )

        st.dataframe(
            flagged,
            use_container_width=True
        )

        # =====================================================
        # EXPORT EXCEL
        # =====================================================

        with tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".xlsx"
        ) as tmp:

            temp_path = tmp.name

        with pd.ExcelWriter(
            temp_path,
            engine="openpyxl"
        ) as writer:

            summary_df.to_excel(
                writer,
                sheet_name="Summary",
                index=False
            )

            flagged.to_excel(
                writer,
                sheet_name="UID Mismatches",
                index=False
            )

        # =====================================================
        # FORMAT EXCEL
        # =====================================================

        wb = load_workbook(temp_path)

        bold_font = Font(bold=True)

        for sheet_name in wb.sheetnames:

            ws = wb[sheet_name]

            # HEADER STYLE
            for cell in ws[1]:
                cell.font = bold_font

            # FREEZE HEADER
            ws.freeze_panes = "A2"

            # FILTERS
            ws.auto_filter.ref = ws.dimensions

            # AUTO WIDTH
            for col in ws.columns:

                max_length = 0
                column = col[0].column

                for cell in col:

                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(
                    max_length + 2,
                    40
                )

                ws.column_dimensions[
                    get_column_letter(column)
                ].width = adjusted_width

        wb.save(temp_path)

        # =====================================================
        # DOWNLOAD
        # =====================================================

        with open(temp_path, "rb") as f:
            file_bytes = f.read()

        st.download_button(
            label="⬇ Download Results",
            data=file_bytes,
            file_name="retail_uid_results.xlsx",
            mime=(
                "application/"
                "vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )
